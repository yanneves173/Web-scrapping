import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, fills, Border, Side

#removing request warnings
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

#________getting the url
bbc_url = 'https://www.bbc.com/news'
theguardian_url = 'https://www.theguardian.com/world'
bbc_response = requests.get(bbc_url, verify = False)
theguardian_response = requests.get(theguardian_url, verify = False)
#print(bbc_response)

#________creating a list with all news
bbc_news=[]
theguardian_news = []


"""BBC"""

#________web scrapping from BBC
bbc_soup = BeautifulSoup(bbc_response.text, 'html.parser')
bbc_headlines = bbc_soup.find('body').find_all('h3')
unwanted = ['BBC World News TV', 'BBC World Services Radio', 'News daily newsletter', 'Mobile app', 'Get in touch']
#bbc_datetime = bbc_soup.find('li').find_all('time')
for x in list(dict.fromkeys(bbc_headlines)):
    if x.text.strip() not in unwanted and x.text.strip().__contains__('Weekly quiz')==False:
        bbc_news.append(x.text.strip())
#print(bbc_news)

"""THE GUARDIAN"""

#________web scrapping from THE GUARDIAN
theguardian_soup = BeautifulSoup(theguardian_response.text, 'html.parser')
theguardian_headlines = theguardian_soup.find('body').find_all('h3')
unwanted = []

for x in list(dict.fromkeys(theguardian_headlines)):
    if x.text.strip() not in unwanted and x.text.strip().__contains__('Weekly quiz')==False:
        theguardian_news.append(x.text.strip())

news=bbc_news + theguardian_news

### Creating a database



    
def Create_database(news):

    wb = Workbook()
    sheet = wb.worksheets[0]
    ws = wb["Sheet"]

    #Header

    list_header = ["Title","Source", "date"]

    for index, header in enumerate(list_header):
        sheet[f'{chr(ord("A")+ index)}1'] = header
        font_style = Font(size=12,bold=True)
        sheet[f'{chr(ord("A")+ index)}1'].font = font_style
        bottom = Side(border_style="thin", color="000000")
        border= Border(bottom=bottom)
        sheet[f'{chr(ord("A")+ index)}1'].border=border

    #Inputing -news- on excel file -- Column 1
    
    for i in range(0,len(news)):
        ws.cell(row=2+i, column=1).value = news[i]

    #Inputing -Source- on excel file -- Column 2
    for i in range(0, len(news)):
        if news[i] in bbc_news:
            ws.cell(row=2+i, column = 2).value = "BBC News"
        elif news[i] in theguardian_news:
            ws.cell(row=2+i, column = 2).value = "The Guardian"
    wb.save("news.xlsx")


Create_database(news)
