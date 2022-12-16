# import sys
# from cx_Freeze import setup, Executable
from tkinter.font import BOLD
import PySimpleGUI as sg
from typing import Pattern
import win32com.client
import win32timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Color, PatternFill, fills, Border, Side
# build_exe_options = {"packages": ["os"], "includes": ["PysimpleGUI", "typing", "win32com.client", "win32timezone", "openpyxl","openpyxl.styles"], "include_files": ["ss.ico"]}
# base = None
# if sys.platform =="win32":
#     base = "Win32GUI"

# setup(
#     name="TARSEXECUT",
#     version="1.0"
#     description = "getting our travel alert reports data",
#     options=("build_exe": build_exe_options),
#     executables = [Executable(script="TARSEXECUT.py", base=base, icon="ss.ico")]
#     )
    
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_color(ws, cell_range, Color):
    Pattern_fill = fills.PatternFill("solid", fgColor= Color)
    for row in ws[cell_range]:
        for cell in row:
            cell.fill= Pattern_fill


def tar_report(tst=0):
    outlook = win32com.client.Dispatch("outlook.application")
    mapi = outlook.GetNamespace("MAPI")
    account = [i for i in mapi.Accounts][0]
    inbox = mapi.GetDefaultFolder(6)
    test_folder = inbox.Folders["TravelAlertReport"]
    items = test_folder.Items

    wb= Workbook()
    planilha = wb.worksheets[0]
    ws=wb["Sheet"]

    #Header

    list_header = ["Status", "Level", "Location", "Category", "Date", "Month", "Year","Casualities?(Deaths/Injuries)", "Relevant?(Yes/No/Update)", "Justification", "Used by TST", "Learning Required?"]

    for index, header in enumerate(list_header):
        planilha[f'{chr(ord("A")+ index)}1'] = header
        font_style = Font(size=12,bold=True)
        planilha[f'{chr(ord("A")+ index)}1'].font = font_style
        bottom = Side(border_style="thin", color="000000")
        border= Border(bottom=bottom)
        planilha[f'{chr(ord("A")+ index)}1'].border=border

    #filter

    ws.auto_filter.ref= "A1:K99999"
    ws.auto_filter.add_filter_column(1,["Status", "Level", "Location", "Category", "Date", "Month", "Year","Casualities?(Deaths/Injuries)", "Relevant?(Yes/No/Update)", "Justification", "Used by TST", "Learning Required?"] )

    #Column width dimensions
    list_dimensions = ["18", "22.33", "28", "62.28", "20", "6.56", "4.73", "19.67", "19.67", "55", "11.44", "18.33"]

    for index, dimensions in enumerate(list_dimensions):
        ws.column_dimensions[f'{chr(ord("A")+ index)}'].width = dimensions
        
        

    count_isos=0
    count_escalated = 0
    for index,item in enumerate(items):
        
        if "RE:" in item.Subject or "FW:" in item.subject:
            count_isos+=1 
            # Status - RTA
            ws.cell(row=index+2, column=1).value= "Escalated"
            count_escalated +=1

            #Level - Replied Travel Alerts(RTA)
            incident_index = item.body.index("incident") + 12
            category_indexpos = item.body[incident_index:].index("Category") + incident_index
            level = item.body[incident_index:category_indexpos]
            ws.cell(row=index+2, column = 2).value = level
            
            if "Advisory" in level:
                # advisory_pattern = Pattern(patternType = "solid", fgColor="FFFF00")
                ws.cell(row=index+2, column = 2).fill = fills.PatternFill("solid", fgColor="FFFF00")
            
            if "Notice" in level:
                # notice_pattern = Pattern(patternType = "solid", fgColor = "00B050")
                ws.cell(row=index+2, column = 2).fill = fills.PatternFill("solid", fgColor = "00B050")
            if "Special Advisory" in level or "Evacuation" in level:
                # special_pattern = Pattern(patternType = "solid", fgColor= "FF0000")
                ws.cell(row=index+2, column = 2).fill = fills.PatternFill("solid", fgColor= "F20000")
            
            #location - RTA
            location_index = item.body.index("Location") +12
            time_indexpos = item.body[location_index:].index("Time")+ location_index
            location = item.body[location_index:time_indexpos]
            ws.cell(row=index+2, column = 3).value = location

            #category - RTA
            category_index = item.body.index("Category") + 12
            location_indexpos = item.body[category_index:].index("Location") + category_index
            category = item.body[category_index:location_indexpos]
            ws.cell(row=index+2, column = 4).value = category

            #Time
            date_time = item.ReceivedTime.replace(tzinfo=None)
            ws.cell(row=index+2, column = 5).value = date_time

            #Month
            date_month = item.ReceivedTime.month
            ws.cell(row=index+2, column = 6).value = date_month

            #Year
            date_year = item.ReceivedTime.year
            ws.cell(row=index+2, column = 7).value = date_year
            
            # #Casualities?(Deaths/Injuries) - RTA
            # reported_index = item.body.index("reported") + 12
            # potential_indexpos = item.body[reported_index:].index("Potential ") + reported_index
            # casualities = item.body[reported_index:potential_indexpos]
            # ws.cell(row=index+2, column =8).value = casualities

            #Relevant?(Yes/No/Update) - RTA
            assessed_index = item.body.index("assessed") + 12
            additional_indexpos = item.body[assessed_index:].index("Additional") + assessed_index
            assessed = item.body[assessed_index:additional_indexpos]
            ws.cell(row=index+2, column = 9).value = assessed
        
            #comments - RTA
            just_index = item.body.index("Justification") + 17
            in_indexpos = item.body[just_index:].index("In") + just_index
            comments = item.body[just_index:in_indexpos]
            ws.cell(row=index+2, column = 10).value = comments
            
            #Used by TST
            ws.cell(row=index+2, column=11).value= "No"
            
        
        else:
            count_isos+=1
            #Status
            ws.cell(row=index+2, column = 1).value = "No Action Taken"

            #Level - Travel Alerts (TA)
            level_index = item.body.index("Level:") + 8
            https_indexpos = item.body[level_index:].index("<https:") + level_index
            levelta = item.body[level_index:https_indexpos]
            ws.cell(row=index+2, column = 2).value = levelta

            if "Advisory" in levelta:
                ws.cell(row=index+2, column = 2).fill = fills.PatternFill("solid", fgColor="FFFF00")
            
            if "Notice" in levelta:
                ws.cell(row=index+2, column = 2).fill = fills.PatternFill("solid", fgColor = "00B050")

            if "Special Advisory" in levelta or "Evacuation" in levelta:
                ws.cell(row=index+2, column = 2).fill = fills.PatternFill("solid", fgColor= "F20000")
            
            #location - TA
            location_index = item.body.index("Location:") + 11
            https_indexposloc = item.body[location_index:].index("<https:")+ location_index
            locationta = item.body[location_index:https_indexposloc]
            ws.cell(row=index+2, column = 3).value = locationta

            #category - TA
            category_index = item.body.index("Category:") + 11
            https_indexpos = item.body[category_index:].index("<https:") + category_index
            categoryta = item.body[category_index:https_indexpos]
            ws.cell(row=index+2, column = 4).value = categoryta
        
            #Time
            date_time = item.ReceivedTime.replace(tzinfo=None)
            ws.cell(row=index+2, column = 5).value = date_time
            
            #Month
            date_month = item.ReceivedTime.month
            ws.cell(row=index+2, column = 6).value = date_month

            #Year
            date_year = item.ReceivedTime.year
            ws.cell(row=index+2, column = 7).value = date_year
            
            #Casualities?(Deaths/Injuries) - TA
            ws.cell(row=index+2, column=8).value= "NA"
            
            #Relevant?(Yes/No/Update)
            ws.cell(row=index+2, column = 9).value = "No"
                        
            #Used by TST - TA
            ws.cell(row=index+2, column=11).value= "-"


    #_________________________________________ SUMMARY_______________________________________________________

    planilha1 = wb.create_sheet("Summary", 1)
    ws1 = wb["Summary"]

    #creating the table headers__________________________________________________________________________________

    listx=["Total ISOS received", "Total escalated by EWMR team to TST", "Total used by TST", "Travel Alert Usage (TAU) rate", "Number of relevant cases not escalated to TST"]

    for i in range(0,5):

        ws1.cell(row=i+2, column = 1).value = listx[i]

    monthslist=["January","February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]
    month_to_name= {month_number: month_name for month_number, month_name in zip(range(1,13), monthslist)}

    #arranging the month of the data -- (kind of listY)

    summary_month = month_to_name[items[1].ReceivedTime.replace(tzinfo=None).month]

    ws1.cell(row=1, column=2).value = summary_month

    list_ws1_dimensions = ["40","23"]

    for index, dimensions in enumerate(list_ws1_dimensions):
        ws1.column_dimensions[f'{chr(ord("A")+ index)}'].width = dimensions

    #designing of the table

    set_border(ws1,"A1:B6")

    set_color(ws1, "A1:B6", "538DD5")
    set_color(ws1, "B2:B6", "C5D9F1")
    set_color(ws1, "A1:A1", "000000")

    #Data info of the table
        #Total ISOS received
    ws1.cell(row=2,column=2).value= count_isos

        #Total escalated to TST
    for element in ws["A"]:
        if  "Escalated" == element.value:
            ws1.cell(row=3, column=2).value = count_escalated

    #drawing a chart
    # values = Reference(ws, min_col=2, min_row= 1, max_col = item[11], max_row = items[count_isos] )
    # chart= LineChart()
    

    wb.save("Travel Alert Report.xlsx")


class PythonScreen:
    def __init__(self):
        
        sg.theme('LightBlue1')
        #Layout
        layout = [
            [sg.Image("C:/Users/320144301/Desktop/Travel Alert Project/tar/ss.png",expand_x=True,)],
            [sg.Text("Before generate Travel Alert Report, please follow the following steps:",auto_size_text=True, justification="center",font=("Arial",13) )], 
            [sg.Text("\n \n 1- Create in your outlook inbox a folder named 'TravelAlertReport' ", auto_size_text=True, justification="center", font=("Arial",11))],
            [sg.Text("\n 2- Select all alert mails you want to gather in your report, and then copy them to 'TravelAlertReport' folder in your inbox", auto_size_text=True, justification="center",font=("Arial",11))],
            [sg.Text("\n 3- Be sure that you are selecting only 'No action taken' and 'Escalated' alerts from their respective month folder", auto_size_text=True, justification="center", font=("Arial",11))],
            [sg.Text("\n \n ISOS travel alert", justification= "center", font=("Arial",10))],
            [sg.Button("Generate report (.xlsx)", mouseover_colors="Gray", use_ttk_buttons=True)]
                    ]
        
        #window
        window = sg.Window("Travel Alert Report").layout(layout)
        # extracting data
        self.button, self.values = window.Read()
    
    def Play(self):
        # amount_ta_used = int(self.values[0])
        tar_report()
        # print(self.values)

screen = PythonScreen()
screen.Play()
